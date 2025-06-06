import pandas as pd
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
import subprocess
import time

total = 0.00
total_cheque = 0.00
total_dinheiro = 0.00
total_recibos = 0
cheques = []
recibos_detalhes = []

def atualizacao():
    caminho_py = r"C:\Users\perer\PycharmProjects\fechamento\main.py"
    arquivo_py = os.path.join(os.path.dirname(os.path.abspath(__file__)), caminho_py)
    arquivo_exe = r"C:\Users\perer\PycharmProjects\fechamento\dist\main.exe"

    data_py = os.path.getmtime(arquivo_py)

    if not os.path.exists(arquivo_exe):
        data_exe = 0
        
    else:
        data_exe = os.path.getmtime(arquivo_exe)

    if data_py > data_exe:
        print("Saiu nova atualização! atualizando...")
        subprocess.run(["pyinstaller", "--onefile", "main.py"])

        while not os.path.exists(arquivo_exe):
            print("⌛ Aguarde, gerando o novo executável...")
            time.sleep(2)

        print("✅ Atualização concluída! Reiniciando...")
        os.system(f"start {arquivo_exe}")

    else:
        print("Seu sistema está atualizado!")

def recibos():
    global total, total_recibos
    while True:
        recibo_input = input("Digite o Valor do recibo (Ou pressione ENTER para finalizar): ")
        if recibo_input == "":
            print(f"O valor total do caixa é de {total}")
            break
        
        try:
            recibo = float(recibo_input)
            qtd = int(input("Quantidade de recibos: "))

        except ValueError:
            print("X Entrada inválida! Digite apenas números!")
            continue

        subtotal = recibo * qtd
        total += subtotal
        total_recibos += qtd
        recibos_detalhes.append({"Quantidade": qtd, "Valor do Recibo": recibo, "Valor Total": subtotal})
   
    return recibos_detalhes

def fechamento():
    global total_cheque, total_dinheiro
    while True:
        cheques_input = input('Digite o valor do cheque separados por virgula: ')
        if cheques_input:
            try:
                valores_cheques = [float(valor.strip()) for valor in cheques_input.split(",")]
                cheques.extend(valores_cheques)
                total_cheque = sum(cheques)
                subtotal_cheque = total - total_cheque
                print(f"O valor total - cheques é de R${subtotal_cheque}.")
                break
            
            except ValueError:
                print("X Entrada inválida! Certifique-se de separar corretamente por vírgula!")
        else:
            break
        
            

    # Somar valores de dinheiro até o usuário pressionar Enter sem digitar nada pra finalizar
    print('Digite um valor em dinheiro um por vez, ou pressione ENTER sem nada para finalizar: ')
    while True:
        dinheiro_input = input('Valor em dinheiro: ')
        if dinheiro_input == "":
            break
        try:
            total_dinheiro += float(dinheiro_input)
            subtotal_geral = subtotal_cheque - total_dinheiro
            print(f"O subtotal geral é de R${subtotal_geral}")

        except ValueError:
            print("X Entrada Inválida! Digite apenas números.")

def salvar_fechamento():
    arquivo_excel = os.path.join(os.getcwd(), "fechamento_caixa.xlsx")
    data_fechamento = f"Fechamento_{datetime.now().strftime('%d-%m-%Y')}"

    df_recibos = pd.DataFrame(recibos_detalhes)
    
    totais_formatados = [
        {"Quantidade": "", "Valor do Recibo": "Total Recibos", "Valor Total": total},
        {"Quantidade": "", "Valor do Recibo": "Total Cheques", "Valor Total": total_cheque},
        {"Quantidade": "", "Valor do Recibo": "Total Dinheiro", "Valor Total": total_dinheiro},
    ]

    df_totais = pd.DataFrame(totais_formatados)

    df_final = pd.concat([df_recibos, pd.DataFrame([{}]), df_totais], ignore_index=True)

    if not os.path.exists(arquivo_excel):
        wb = Workbook()
        ws = wb.active
        ws.title = "Fechamentos"
        wb.save(arquivo_excel)
        print(f"Arquivo '{arquivo_excel}' criado porque não existia.")

    if os.path.exists(arquivo_excel):
        wb = load_workbook(arquivo_excel)
        if data_fechamento in wb.sheetnames:
            del wb[data_fechamento]
            wb.save(arquivo_excel)

    with pd.ExcelWriter(arquivo_excel, engine="openpyxl", mode='a') as writer:
        df_final.to_excel(writer, sheet_name=data_fechamento, index=False)

    print(f"✅ Fechamento de Caixa Salvo na aba '{data_fechamento}' dentro do arquivo {arquivo_excel}!")
    input("Pressione ENTER para sair ...")

atualizacao()
recibos()
fechamento()
salvar_fechamento()